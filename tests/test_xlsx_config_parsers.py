from __future__ import annotations

import tempfile
from pathlib import Path
from unittest import TestCase

from openpyxl import Workbook

from infrastructure.xlsx_region_parser import (
    ImportRuleError,
    load_restricted_regions_from_xlsx,
)
from infrastructure.xlsx_sku_group_parser import load_sku_groups_from_xlsx


def _make_region_xlsx(rows: list[list[str | None]], merges: list[str] | None = None) -> Path:
    """Helper: create a region xlsx with given rows and optional merges."""
    tmp_dir = tempfile.mkdtemp()
    path = Path(tmp_dir) / "regions.xlsx"
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    if merges:
        for merge_range in merges:
            ws.merge_cells(merge_range)
    wb.save(path)
    wb.close()
    return path


class RegionParserLegalTests(TestCase):
    """合法用例：必须正确解析，不报错。"""

    def test_single_row_sku_province(self) -> None:
        """单行 SKU 限发省（市为空）。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A", "新疆", None],
        ])
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(parsed, [{"sku_code": "SKU-A", "province": "新疆", "city": None}])

    def test_same_sku_multiple_rows(self) -> None:
        """同 SKU 多行多区域。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A", "新疆", None],
            ["SKU-A", "西藏", "拉萨"],
            ["SKU-A", "四川", "甘孜"],
        ])
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(parsed, [
            {"sku_code": "SKU-A", "province": "新疆", "city": None},
            {"sku_code": "SKU-A", "province": "西藏", "city": "拉萨"},
            {"sku_code": "SKU-A", "province": "四川", "city": "甘孜"},
        ])

    def test_real_merged_sku_column(self) -> None:
        """SKU 列真实纵向合并。"""
        path = _make_region_xlsx(
            [
                ["产品名称", "省", "市"],
                ["SKU-A", "新疆", None],
                [None, "西藏", "拉萨"],
                [None, "四川", "甘孜"],
            ],
            merges=["A2:A4"],
        )
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(parsed, [
            {"sku_code": "SKU-A", "province": "新疆", "city": None},
            {"sku_code": "SKU-A", "province": "西藏", "city": "拉萨"},
            {"sku_code": "SKU-A", "province": "四川", "city": "甘孜"},
        ])

    def test_multi_sku_in_one_cell(self) -> None:
        """一个单元格多个 SKU（换行分隔）。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A\nSKU-B", "新疆", None],
        ])
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(parsed, [
            {"sku_code": "SKU-A", "province": "新疆", "city": None},
            {"sku_code": "SKU-B", "province": "新疆", "city": None},
        ])

    def test_comma_in_product_name_preserved(self) -> None:
        """逗号出现在产品名称内部时不应被拆分。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["酸菜鱼450g/450g*20袋（一件代发，单次不低于5袋)", "青海", None],
        ])
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(len(parsed), 1)
        self.assertEqual(parsed[0]["sku_code"], "酸菜鱼450g/450g*20袋（一件代发，单次不低于5袋)")

    def test_semicolon_in_product_name_preserved(self) -> None:
        """分号出现在产品名称内部时不应被拆分。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A；SKU-B", "新疆", None],
        ])
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(len(parsed), 1)
        self.assertEqual(parsed[0]["sku_code"], "SKU-A；SKU-B")

    def test_fullwidth_comma_in_product_name_preserved(self) -> None:
        """全角逗号出现在产品名称内部时不应被拆分。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["大别山老母鸡1kg*10只/箱/（5箱以上，下此链接）", "青海", None],
        ])
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(len(parsed), 1)
        self.assertEqual(parsed[0]["sku_code"], "大别山老母鸡1kg*10只/箱/（5箱以上，下此链接）")

    def test_multi_sku_plus_merged_region(self) -> None:
        """多个 SKU + 合并区域：每行的 SKU 与区域按行展开。"""
        path = _make_region_xlsx(
            [
                ["产品名称", "省", "市"],
                ["SKU-A\nSKU-B", "新疆", None],
                [None, "西藏", None],
            ],
            merges=["A2:A3"],
        )
        parsed = load_restricted_regions_from_xlsx(path)
        # 合并区域逐行展开，每行内 SKU 也展开 → 行优先顺序
        self.assertEqual(parsed, [
            {"sku_code": "SKU-A", "province": "新疆", "city": None},
            {"sku_code": "SKU-B", "province": "新疆", "city": None},
            {"sku_code": "SKU-A", "province": "西藏", "city": None},
            {"sku_code": "SKU-B", "province": "西藏", "city": None},
        ])

    def test_exact_duplicate_auto_dedup(self) -> None:
        """完全重复规则自动去重。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A", "新疆", None],
            ["SKU-A", "新疆", None],
        ])
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(parsed, [{"sku_code": "SKU-A", "province": "新疆", "city": None}])

    def test_province_override_city(self) -> None:
        """省级限发覆盖市级限发。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A", "新疆", None],
            ["SKU-A", "新疆", "乌鲁木齐"],
        ])
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(parsed, [{"sku_code": "SKU-A", "province": "新疆", "city": None}])

    def test_province_override_city_multiple_skus(self) -> None:
        """省级限发只覆盖同 SKU 的市级，不影响其他 SKU。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A", "新疆", None],
            ["SKU-A", "新疆", "乌鲁木齐"],
            ["SKU-B", "新疆", "乌鲁木齐"],
        ])
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(parsed, [
            {"sku_code": "SKU-A", "province": "新疆", "city": None},
            {"sku_code": "SKU-B", "province": "新疆", "city": "乌鲁木齐"},
        ])

    def test_city_remark_filtered(self) -> None:
        """市列中的备注文字被过滤。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A", "新疆", "新增"],
        ])
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(parsed, [{"sku_code": "SKU-A", "province": "新疆", "city": None}])

    def test_empty_rows_skipped(self) -> None:
        """整行空行被跳过。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A", "新疆", None],
            [None, None, None],
            ["SKU-B", "西藏", None],
        ])
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(parsed, [
            {"sku_code": "SKU-A", "province": "新疆", "city": None},
            {"sku_code": "SKU-B", "province": "西藏", "city": None},
        ])


class RegionParserIllegalTests(TestCase):
    """非法用例：必须立即报错。"""

    def test_empty_sku_not_in_merge_raises(self) -> None:
        """普通空白 SKU 行，不在合并单元格内，必须报错。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A", "新疆", None],
            [None, "西藏", None],
        ])
        with self.assertRaises(ImportRuleError) as ctx:
            load_restricted_regions_from_xlsx(path)
        self.assertEqual(ctx.exception.row, 3)
        self.assertEqual(ctx.exception.column, "A")

    def test_merged_sku_first_cell_empty_raises(self) -> None:
        """SKU 合并单元格首格为空，必须报错。"""
        path = _make_region_xlsx(
            [
                ["产品名称", "省", "市"],
                [None, "新疆", None],
                [None, "西藏", None],
            ],
            merges=["A2:A3"],
        )
        with self.assertRaises(ImportRuleError) as ctx:
            load_restricted_regions_from_xlsx(path)
        self.assertEqual(ctx.exception.row, 2)
        self.assertIn("首格为空", ctx.exception.reason)

    def test_province_empty_but_city_has_value_raises(self) -> None:
        """省为空但市/区有值，必须报错。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A", None, "乌鲁木齐"],
        ])
        with self.assertRaises(ImportRuleError) as ctx:
            load_restricted_regions_from_xlsx(path)
        self.assertEqual(ctx.exception.row, 2)
        self.assertEqual(ctx.exception.column, "B")
        self.assertIn("省为空但市/区有值", ctx.exception.reason)

    def test_province_empty_raises(self) -> None:
        """省为空（市也为空），应报错。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A", None, None],
        ])
        with self.assertRaises(ImportRuleError) as ctx:
            load_restricted_regions_from_xlsx(path)
        self.assertEqual(ctx.exception.row, 2)
        self.assertEqual(ctx.exception.column, "B")
        self.assertIn("省和市均为空", ctx.exception.reason)

    def test_header_missing_raises(self) -> None:
        """表头缺失，必须报错。"""
        path = _make_region_xlsx([
            ["产品", "省份", "城市"],  # 不符合要求的表头
            ["SKU-A", "新疆", None],
        ])
        with self.assertRaises(ImportRuleError) as ctx:
            load_restricted_regions_from_xlsx(path)
        self.assertEqual(ctx.exception.row, 1)
        self.assertIn("表头", ctx.exception.reason)

    def test_no_data_rows_raises(self) -> None:
        """只有表头无数据行。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
        ])
        with self.assertRaises(ImportRuleError):
            load_restricted_regions_from_xlsx(path)

    def test_sku_merge_crosses_columns_raises(self) -> None:
        """SKU 合并单元格横跨多列，必须报错。"""
        path = _make_region_xlsx(
            [
                ["产品名称", "省", "市"],
                ["SKU-A", "新疆", None],
            ],
            merges=["A2:B2"],
        )
        with self.assertRaises(ImportRuleError) as ctx:
            load_restricted_regions_from_xlsx(path)
        self.assertIn("跨越", ctx.exception.reason)

    def test_newline_in_sku_cell_is_multi_sku(self) -> None:
        """单元格内含换行符作为 multi-SKU 分隔符，合法。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A\nSKU-B", "新疆", None],
        ])
        parsed = load_restricted_regions_from_xlsx(path)
        self.assertEqual(parsed, [
            {"sku_code": "SKU-A", "province": "新疆", "city": None},
            {"sku_code": "SKU-B", "province": "新疆", "city": None},
        ])

    def test_newline_in_province_cell_raises(self) -> None:
        """省列单元格内含换行符，必须报错。"""
        path = _make_region_xlsx([
            ["产品名称", "省", "市"],
            ["SKU-A", "新疆\n西藏", None],
        ])
        with self.assertRaises(ImportRuleError) as ctx:
            load_restricted_regions_from_xlsx(path)
        self.assertEqual(ctx.exception.column, "B")


class SkuGroupParserTests(TestCase):
    """SKU 群配置解析器测试（保留原有逻辑）。"""

    def test_sku_group_parser_expands_merged_product_block_once(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "sku_groups.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["产品名称", "厂家群名", "群主手机号"])
            ws["A2"] = (
                "蚕丝丝光罩宽肩带内衣/月光白/XL(115-140)"
                "        "
                "男士抗菌棉质内裤(4条装)/颜色随机4XL(175-200)\n"
                "云感棉朵蚕丝内裤(高腰4条装)/颜色随机  L(105-125)"
            )
            ws["B2"] = "婷美&袜子+内衣沟通群"
            ws["C2"] = "18231132648"
            ws.merge_cells("A2:A4")
            ws.merge_cells("B2:B4")
            ws.merge_cells("C2:C4")
            wb.save(path)
            wb.close()

            parsed = load_sku_groups_from_xlsx(path)

        self.assertEqual(
            parsed,
            [
                {
                    "sku_code": "蚕丝丝光罩宽肩带内衣/月光白/XL(115-140)",
                    "group_name": "婷美&袜子+内衣沟通群",
                    "owner_mobile": "18231132648",
                },
                {
                    "sku_code": "男士抗菌棉质内裤(4条装)/颜色随机4XL(175-200)",
                    "group_name": "婷美&袜子+内衣沟通群",
                    "owner_mobile": "18231132648",
                },
                {
                    "sku_code": "云感棉朵蚕丝内裤(高腰4条装)/颜色随机  L(105-125)",
                    "group_name": "婷美&袜子+内衣沟通群",
                    "owner_mobile": "18231132648",
                },
            ],
        )
