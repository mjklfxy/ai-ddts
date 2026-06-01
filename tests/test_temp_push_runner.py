"""Pure-function unit tests for temp push rule chain, order splitting, and Excel output."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from unittest import TestCase

from application.file_generator import ExcelFileGenerator, ORDER_ERROR_HEADERS
from application.manual_runner import summarize_result
from application.order_splitter import GroupOrderBatch, OrderLineForSplit, OrderSplitter
from application.pipeline import (
    PipelineBatchDelivery,
    PipelineOrderEvaluation,
    PipelineRunResult,
)
from application.task_service import TaskContext
from domain.enums.rule import RuleDecision
from domain.enums.status import KingdeeStatus, PushStatus
from domain.exception_order import ExceptionOrder
from domain.rule_engine import RuleEngine
from domain.rules.base import RuleContext
from domain.rules.group_rule import GroupRule
from domain.rules.region_rule import RegionRule, RestrictedRegion
from domain.rules.special_sku_rule import SpecialSkuRule
from domain.sku_group_info import SkuGroupInfo
from infrastructure.message_adapter import MessageSendResult


# ---------------------------------------------------------------------------
# SpecialSkuRule
# ---------------------------------------------------------------------------


class SpecialSkuRuleTests(TestCase):
    """Tests SpecialSkuRule positive-selection logic."""

    def test_sku_in_list_returns_pass(self) -> None:
        rule = SpecialSkuRule(special_skus={"SKU-001", "SKU-002"})
        ctx = RuleContext(order_no="SO-1", sku_codes=("SKU-001",))
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.PASS)

    def test_sku_not_in_list_returns_ignore(self) -> None:
        rule = SpecialSkuRule(special_skus={"SKU-001"})
        ctx = RuleContext(order_no="SO-1", sku_codes=("SKU-999",))
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.IGNORE)

    def test_empty_sku_list_returns_ignore(self) -> None:
        rule = SpecialSkuRule(special_skus=set())
        ctx = RuleContext(order_no="SO-1", sku_codes=("SKU-001",))
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.IGNORE)

    def test_multiple_skus_any_match_passes(self) -> None:
        rule = SpecialSkuRule(special_skus={"SKU-001"})
        ctx = RuleContext(order_no="SO-1", sku_codes=("SKU-999", "SKU-001"))
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.PASS)


# ---------------------------------------------------------------------------
# RegionRule
# ---------------------------------------------------------------------------


class RegionRuleTests(TestCase):
    """Tests RegionRule restricted-region matching."""

    def test_sku_and_province_match_returns_error(self) -> None:
        rule = RegionRule(restricted_regions=[
            RestrictedRegion(sku_code="SKU-001", province="广东省"),
        ])
        ctx = RuleContext(
            order_no="SO-1",
            sku_codes=("SKU-001",),
            receiver_province="广东省",
            receiver_city="深圳市",
        )
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.ERROR)
        self.assertEqual(result.rule_name, "RegionRule")

    def test_sku_not_in_restricted_list_returns_pass(self) -> None:
        rule = RegionRule(restricted_regions=[
            RestrictedRegion(sku_code="SKU-001", province="广东省"),
        ])
        ctx = RuleContext(
            order_no="SO-1",
            sku_codes=("SKU-999",),
            receiver_province="广东省",
        )
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.PASS)

    def test_province_mismatch_returns_pass(self) -> None:
        rule = RegionRule(restricted_regions=[
            RestrictedRegion(sku_code="SKU-001", province="广东省"),
        ])
        ctx = RuleContext(
            order_no="SO-1",
            sku_codes=("SKU-001",),
            receiver_province="浙江省",
        )
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.PASS)

    def test_disabled_rule_returns_pass(self) -> None:
        rule = RegionRule(
            restricted_regions=[RestrictedRegion(sku_code="SKU-001", province="广东省")],
            enabled=False,
        )
        ctx = RuleContext(
            order_no="SO-1",
            sku_codes=("SKU-001",),
            receiver_province="广东省",
        )
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.PASS)

    def test_city_match_returns_error(self) -> None:
        rule = RegionRule(restricted_regions=[
            RestrictedRegion(sku_code="SKU-001", province="广东省", city="深圳市"),
        ])
        ctx = RuleContext(
            order_no="SO-1",
            sku_codes=("SKU-001",),
            receiver_province="广东省",
            receiver_city="深圳市",
        )
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.ERROR)

    def test_city_mismatch_returns_pass(self) -> None:
        rule = RegionRule(restricted_regions=[
            RestrictedRegion(sku_code="SKU-001", province="广东省", city="深圳市"),
        ])
        ctx = RuleContext(
            order_no="SO-1",
            sku_codes=("SKU-001",),
            receiver_province="广东省",
            receiver_city="广州市",
        )
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.PASS)


# ---------------------------------------------------------------------------
# GroupRule
# ---------------------------------------------------------------------------


class GroupRuleTests(TestCase):
    """Tests GroupRule push-group configuration check."""

    def test_sku_not_in_map_returns_error(self) -> None:
        rule = GroupRule(sku_group_map={
            "SKU-001": SkuGroupInfo(group_name="G1", owner_mobile=""),
        })
        ctx = RuleContext(order_no="SO-1", sku_codes=("SKU-999",))
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.ERROR)
        self.assertEqual(result.rule_name, "GroupRule")

    def test_sku_in_map_returns_pass(self) -> None:
        rule = GroupRule(sku_group_map={
            "SKU-001": SkuGroupInfo(group_name="G1", owner_mobile=""),
        })
        ctx = RuleContext(order_no="SO-1", sku_codes=("SKU-001",))
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.PASS)

    def test_disabled_rule_returns_pass(self) -> None:
        rule = GroupRule(sku_group_map={}, enabled=False)
        ctx = RuleContext(order_no="SO-1", sku_codes=("SKU-001",))
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.PASS)

    def test_empty_group_name_filtered_out(self) -> None:
        rule = GroupRule(sku_group_map={
            "SKU-001": SkuGroupInfo(group_name="", owner_mobile=""),
        })
        ctx = RuleContext(order_no="SO-1", sku_codes=("SKU-001",))
        result = rule.evaluate(ctx)
        self.assertEqual(result.decision, RuleDecision.ERROR)


# ---------------------------------------------------------------------------
# RuleEngine short-circuit
# ---------------------------------------------------------------------------


class RuleEngineTests(TestCase):
    """Tests RuleEngine evaluation order and short-circuit behavior."""

    def test_first_error_stops_chain(self) -> None:
        """When SpecialSkuRule IGNOREs, RegionRule and GroupRule should not run."""
        engine = RuleEngine(rules=[
            SpecialSkuRule(special_skus={"SKU-001"}),
            RegionRule(restricted_regions=[
                RestrictedRegion(sku_code="SKU-001", province="广东省"),
            ]),
            GroupRule(sku_group_map={
                "SKU-001": SkuGroupInfo(group_name="G1", owner_mobile=""),
            }),
        ])
        ctx = RuleContext(
            order_no="SO-1",
            sku_codes=("SKU-999",),
            receiver_province="广东省",
        )
        result = engine.evaluate(ctx)
        self.assertTrue(result.is_ignore)
        # Only SpecialSkuRule should have produced a result
        self.assertEqual(len(result.results), 1)
        self.assertEqual(result.results[0].rule_name, "SpecialSku")

    def test_all_pass_returns_pass(self) -> None:
        engine = RuleEngine(rules=[
            SpecialSkuRule(special_skus={"SKU-001"}),
            RegionRule(restricted_regions=[]),
            GroupRule(sku_group_map={
                "SKU-001": SkuGroupInfo(group_name="G1", owner_mobile=""),
            }),
        ])
        ctx = RuleContext(
            order_no="SO-1",
            sku_codes=("SKU-001",),
            receiver_province="广东省",
        )
        result = engine.evaluate(ctx)
        self.assertTrue(result.is_pass)
        self.assertEqual(len(result.results), 3)

    def test_region_error_before_group(self) -> None:
        """RegionRule ERROR should stop before GroupRule."""
        engine = RuleEngine(rules=[
            SpecialSkuRule(special_skus={"SKU-001"}),
            RegionRule(restricted_regions=[
                RestrictedRegion(sku_code="SKU-001", province="广东省"),
            ]),
            GroupRule(sku_group_map={
                "SKU-001": SkuGroupInfo(group_name="G1", owner_mobile=""),
            }),
        ])
        ctx = RuleContext(
            order_no="SO-1",
            sku_codes=("SKU-001",),
            receiver_province="广东省",
        )
        result = engine.evaluate(ctx)
        self.assertTrue(result.is_error)
        self.assertEqual(result.results[-1].rule_name, "RegionRule")

    def test_temp_push_region_filtering_catches_restricted_sku(self) -> None:
        """Temp push rule chain: SKU in special_skus + region match → RegionRule ERROR."""
        engine = RuleEngine(rules=[
            SpecialSkuRule(special_skus={"SKU-001", "SKU-002"}),
            RegionRule(restricted_regions=[
                RestrictedRegion(sku_code="SKU-001", province="广东省"),
            ], enabled=True),
            GroupRule(sku_group_map={
                "SKU-001": SkuGroupInfo(group_name="G1", owner_mobile=""),
                "SKU-002": SkuGroupInfo(group_name="G2", owner_mobile=""),
            }, enabled=True),
        ])
        # SKU-001 in special_skus + region match → ERROR
        ctx_restricted = RuleContext(
            order_no="SO-1",
            sku_codes=("SKU-001",),
            receiver_province="广东省",
        )
        result = engine.evaluate(ctx_restricted)
        self.assertTrue(result.is_error, "SKU-001 in restricted region should ERROR")
        self.assertEqual(result.results[-1].rule_name, "RegionRule")

        # SKU-002 in special_skus + region NOT match → PASS
        ctx_safe = RuleContext(
            order_no="SO-2",
            sku_codes=("SKU-002",),
            receiver_province="浙江省",
        )
        result = engine.evaluate(ctx_safe)
        self.assertTrue(result.is_pass, "SKU-002 in non-restricted region should PASS")

        # SKU-999 NOT in special_skus → IGNORE (never reaches RegionRule)
        ctx_unknown = RuleContext(
            order_no="SO-3",
            sku_codes=("SKU-999",),
            receiver_province="广东省",
        )
        result = engine.evaluate(ctx_unknown)
        self.assertTrue(result.is_ignore, "SKU-999 not in special_skus should IGNORE")


# ---------------------------------------------------------------------------
# OrderSplitter
# ---------------------------------------------------------------------------


class OrderSplitterTests(TestCase):
    """Tests OrderSplitter group-by logic."""

    def test_same_group_merged_into_one_batch(self) -> None:
        splitter = OrderSplitter(sku_group_map={
            "SKU-001": SkuGroupInfo(group_name="G1", owner_mobile="M1", user_id="U1"),
            "SKU-002": SkuGroupInfo(group_name="G1", owner_mobile="M1", user_id="U1"),
        })
        lines = [
            _make_line("SO-1", "SKU-001"),
            _make_line("SO-2", "SKU-002"),
        ]
        batches = splitter.split(lines)
        self.assertEqual(len(batches), 1)
        self.assertEqual(batches[0].group_name, "G1")
        self.assertEqual(len(batches[0].order_lines), 2)

    def test_different_groups_split_into_batches(self) -> None:
        splitter = OrderSplitter(sku_group_map={
            "SKU-001": SkuGroupInfo(group_name="G1", owner_mobile="M1", user_id="U1"),
            "SKU-002": SkuGroupInfo(group_name="G2", owner_mobile="M2", user_id="U2"),
        })
        lines = [
            _make_line("SO-1", "SKU-001"),
            _make_line("SO-2", "SKU-002"),
        ]
        batches = splitter.split(lines)
        self.assertEqual(len(batches), 2)


# ---------------------------------------------------------------------------
# ExcelFileGenerator.generate_error — field count consistency
# ---------------------------------------------------------------------------


class ExcelErrorFileTests(TestCase):
    """Tests that error Excel rows match header column count."""

    def test_error_row_field_count_matches_header(self) -> None:
        generator = ExcelFileGenerator(
            output_dir=Path("tmp") / "test_error_excel",
            clock=lambda: datetime(2026, 5, 30, 8, 0, 0),
        )
        order = ExceptionOrder(
            order_no="SO-001",
            sku_code="SKU-001",
            delivery_order_no="DO-001",
            goods_summary="Item",
            warehouse_code="WH-001",
            warehouse_name="仓库A",
            quantity=1,
            receiver_name="Receiver",
            address="Address",
            phone="13800000000",
            logistics_company="SF",
            logistics_no="SF-001",
            reason="未配置推送群",
            rule_name="GroupRule",
            channel_classification="直营要货",
        )
        generated = generator.generate_error(
            group_name="G1",
            exception_orders=(order,),
        )
        # Verify the file was created
        self.assertTrue(generated.file_path.exists())
        # Read back and check column count
        from openpyxl import load_workbook
        wb = load_workbook(generated.file_path)
        ws = wb.active
        header_count = len(list(ws.iter_rows(min_row=1, max_row=1))[0])
        data_row = list(ws.iter_rows(min_row=2, max_row=2))[0]
        data_count = len(data_row)
        wb.close()
        self.assertEqual(header_count, data_count,
                         f"Header has {header_count} cols but data has {data_count} cols")
        self.assertEqual(header_count, len(ORDER_ERROR_HEADERS))


# ---------------------------------------------------------------------------
# summarize_result — push status three-state
# ---------------------------------------------------------------------------


class SummarizeResultTests(TestCase):
    """Tests summarize_result push status and count logic."""

    def _make_result(
        self,
        passed: tuple[PipelineOrderEvaluation, ...] = (),
        ignored: tuple[PipelineOrderEvaluation, ...] = (),
        error: tuple[PipelineOrderEvaluation, ...] = (),
        exceptions: tuple[ExceptionOrder, ...] = (),
        deliveries: tuple[PipelineBatchDelivery, ...] = (),
        push_status: PushStatus = PushStatus.PENDING,
    ) -> PipelineRunResult:
        task_context = TaskContext(
            task_id="T-001",
            trace_id="TRACE-001",
            task_name="test",
            created_at=datetime(2026, 5, 30, 8, 0, 0),
            window_start=datetime(2026, 5, 30, 7, 0, 0),
            window_end=datetime(2026, 5, 30, 8, 0, 0),
            push_status=push_status,
            payment_status=None,
            kingdee_status=KingdeeStatus.PENDING,
        )
        return PipelineRunResult(
            task_context=task_context,
            passed_orders=passed,
            ignored_orders=ignored,
            error_orders=error,
            exception_orders=exceptions,
            deliveries=deliveries,
            kingdee_tracking_id=None,
            push_status=push_status,
            kingdee_status=KingdeeStatus.PENDING,
        )

    def test_no_batches_returns_pending(self) -> None:
        result = self._make_result()
        summary = summarize_result(result)
        self.assertEqual(summary.push_status, PushStatus.PENDING)
        self.assertEqual(summary.delivery_count, 0)

    def test_success_status(self) -> None:
        delivery = PipelineBatchDelivery(
            batch=GroupOrderBatch(group_name="G1", owner_mobile="", user_id="U1", order_lines=()),
            generated_file=None,
            message_result=MessageSendResult(
                trace_id="T", group_name="G1", tracking_id="TR-1", attempts=1,
            ),
        )
        eval_ = PipelineOrderEvaluation(order_no="SO-1", engine_result=None)
        result = self._make_result(
            passed=(eval_,),
            deliveries=(delivery,),
            push_status=PushStatus.SUCCESS,
        )
        summary = summarize_result(result)
        self.assertEqual(summary.push_status, PushStatus.SUCCESS)
        self.assertEqual(summary.passed_count, 1)
        self.assertEqual(summary.delivery_count, 1)

    def test_partial_status_with_exception_dedup(self) -> None:
        """Passed orders that also appear in exceptions should be deduped."""
        eval_ = PipelineOrderEvaluation(order_no="SO-1", engine_result=None)
        exception = ExceptionOrder(
            order_no="SO-1", sku_code="SKU-001", delivery_order_no="DO-001",
            goods_summary="Item", warehouse_code="", warehouse_name="",
            quantity=1, receiver_name="R", address="A", phone="P",
            logistics_company="", logistics_no="",
            reason="推送群失败", rule_name="MessagePush",
        )
        result = self._make_result(
            passed=(eval_,),
            exceptions=(exception,),
            push_status=PushStatus.PARTIAL,
        )
        summary = summarize_result(result)
        self.assertEqual(summary.push_status, PushStatus.PARTIAL)
        # SO-1 is in both passed and exception → passed_count = 0
        self.assertEqual(summary.passed_count, 0)
        self.assertEqual(summary.error_count, 1)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_line(order_no: str, sku_code: str) -> OrderLineForSplit:
    return OrderLineForSplit(
        order_no=order_no,
        sku_code=sku_code,
        delivery_order_no=f"DO-{order_no}",
        goods_summary=f"Goods {sku_code}",
        quantity=1,
        receiver_name="Receiver",
        address="Address",
        phone="13800000000",
    )
