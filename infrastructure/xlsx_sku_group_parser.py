from __future__ import annotations

from pathlib import Path
from typing import Any
import re


def load_sku_groups_from_xlsx(xlsx_path: str | Path) -> list[dict[str, str]]:
    """Read SKU group data from an .xls or .xlsx file.

    Column mapping (0-indexed):
      Col 0 — SKU 名称（可能合并多行，一个单元格内多个 SKU 用换行分隔）
      Col 1 — 群名称（可能合并多行）
      Col 2 — 群主手机号（可能合并多行）

    Supports .xls (via xlrd) and .xlsx (via openpyxl).
    """
    path = Path(xlsx_path)
    suffix = path.suffix.lower()

    if suffix == ".xls":
        return _load_via_xlrd(path)
    if suffix in (".xlsx", ".xlsm"):
        return _load_via_openpyxl(path)
    raise ValueError(f"不支持的文件格式: {suffix}")


def _load_via_xlrd(path: Path) -> list[dict[str, str]]:
    import xlrd

    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)

    raw_rows: list[tuple[str, str, str]] = []
    last_sku = ""
    last_group = ""
    last_mobile = ""

    for row_idx in range(1, ws.nrows):
        sku_original = _cell_str_xls(ws, row_idx, 0)
        group_original = _cell_str_xls(ws, row_idx, 1)
        mobile_original = _cell_str_xls(ws, row_idx, 2)

        if not sku_original and not group_original and not mobile_original:
            continue

        sku_text = sku_original or last_sku
        group = group_original or last_group
        mobile = mobile_original or last_mobile

        last_sku = sku_text
        last_group = group
        last_mobile = mobile
        raw_rows.append((sku_text, group, mobile))

    return _expand_skus(raw_rows)


def _load_via_openpyxl(path: Path) -> list[dict[str, str]]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    # === MODIFIED START ===
    # 原因：SKU 群配置 Excel 使用三列表时可能在 A/B/C 列使用合并单元格，逐行 forward-fill 会重复展开，
    #       且复制粘贴产生的超长空白会把两个 SKU 合成一个 key。
    # 影响范围：SKU 群配置 XLSX 上传解析。
    try:
        merged_lookup = _merged_lookup(ws)
        raw_rows: list[tuple[str, str, str]] = []
        seen_product_blocks: set[tuple[int, int]] = set()

        for row_idx in range(2, ws.max_row + 1):
            product_anchor = _merged_anchor(merged_lookup, row_idx, 1)
            if product_anchor in seen_product_blocks:
                continue
            seen_product_blocks.add(product_anchor)

            sku_text = _cell_value(ws, merged_lookup, row_idx, 1)
            if not sku_text:
                continue

            group = _cell_value(ws, merged_lookup, row_idx, 2)
            mobile = _cell_value(ws, merged_lookup, row_idx, 3)
            if not group:
                continue
            raw_rows.append((sku_text, group, mobile))

        return _dedupe_sku_groups(_expand_skus(raw_rows))
    finally:
        wb.close()
    # === MODIFIED END ===


def _cell_str(row: tuple[Any, ...], index: int) -> str:
    if index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def _cell_str_xls(ws, row_idx: int, col_idx: int) -> str:
    try:
        value = ws.cell_value(row_idx, col_idx)
    except IndexError:
        return ""
    if value is None:
        return ""
    return str(value).strip()


def _expand_skus(raw_rows: list[tuple[str, str, str]]) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    for sku_text, group, mobile in raw_rows:
        for product in _split_products(sku_text):
            result.append({
                "sku_code": product,
                "group_name": group,
                "owner_mobile": mobile,
            })
    return result


def _split_products(text: str) -> list[str]:
    """Split products by newline within a cell."""
    # === MODIFIED START ===
    # 原因：Excel 粘贴数据可能用超长空白分隔 SKU；普通空格可能是 SKU 名称的一部分，不能拆。
    # 影响范围：SKU 群配置 XLS/XLSX 上传解析。
    normalized = text.replace("\u00a0", " ").replace("\u3000", " ")
    return [
        part.strip()
        for part in re.split(r"(?:\r\n|\r|\n|\t| {8,})+", normalized)
        if part.strip()
    ]
    # === MODIFIED END ===


# === MODIFIED START ===
# 原因：集中处理 XLSX 合并单元格取值和重复 SKU 冲突校验，避免上传解析生成错误 key。
# 影响范围：SKU 群配置 XLSX 上传解析。
def _merged_lookup(ws) -> dict[tuple[int, int], tuple[int, int, int, int]]:
    lookup: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    for cell_range in ws.merged_cells.ranges:
        bounds = (
            cell_range.min_row,
            cell_range.min_col,
            cell_range.max_row,
            cell_range.max_col,
        )
        for row_idx in range(cell_range.min_row, cell_range.max_row + 1):
            for col_idx in range(cell_range.min_col, cell_range.max_col + 1):
                lookup[(row_idx, col_idx)] = bounds
    return lookup


def _merged_anchor(
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    row_idx: int,
    col_idx: int,
) -> tuple[int, int]:
    bounds = merged_lookup.get((row_idx, col_idx))
    if bounds is None:
        return (row_idx, col_idx)
    return (bounds[0], bounds[1])


def _cell_value(
    ws,
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    row_idx: int,
    col_idx: int,
) -> str:
    anchor_row, anchor_col = _merged_anchor(merged_lookup, row_idx, col_idx)
    value = ws.cell(anchor_row, anchor_col).value
    if value is None:
        return ""
    return str(value).strip()


def _dedupe_sku_groups(items: list[dict[str, str]]) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    seen: dict[str, tuple[str, str]] = {}
    for item in items:
        sku = item["sku_code"]
        value = (item["group_name"], item["owner_mobile"])
        if sku in seen:
            if seen[sku] != value:
                raise ValueError(f"SKU 群配置冲突: {sku}")
            continue
        seen[sku] = value
        result.append(item)
    return result
# === MODIFIED END ===


def load_sku_groups_from_bytes(data: bytes, filename: str = "upload.xlsx") -> list[dict[str, str]]:
    """Parse SKU groups from in-memory xlsx/xls bytes."""
    import tempfile

    suffix = Path(filename).suffix.lower()
    if suffix not in (".xls", ".xlsx", ".xlsm"):
        raise ValueError(f"不支持的文件格式: {suffix}")

    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(data)
        tmp_path = Path(tmp.name)

    try:
        return load_sku_groups_from_xlsx(tmp_path)
    finally:
        tmp_path.unlink(missing_ok=True)
