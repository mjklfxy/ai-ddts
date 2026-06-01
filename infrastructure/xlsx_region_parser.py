"""限发区域 Excel 严格解析器。

规则：
- 每行一条记录，不允许 forward-fill
- 仅支持 A 列（产品名称）真实纵向合并单元格
- 一个单元格支持多个 SKU（换行符分隔）
- 省级限发覆盖市级限发
- 任何格式错误立即中断，返回结构化错误信息
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path


class ImportRuleError(Exception):
    """结构化导入错误，包含行号、列名、原因和建议。"""

    def __init__(self, row: int, column: str, reason: str, suggestion: str) -> None:
        self.row = row
        self.column = column
        self.reason = reason
        self.suggestion = suggestion
        super().__init__(f"第 {row} 行 {column} 列错误：{reason}。{suggestion}")


_CITY_REMARKS = {"新增", "新加", "备注", "新增区域", "新增限发"}

# --- Entry points ---


def load_restricted_regions_from_xlsx(
    xlsx_path: str | Path,
) -> list[dict[str, str | None]]:
    """从 .xls 或 .xlsx 文件严格解析限发区域。"""
    path = Path(xlsx_path)
    suffix = path.suffix.lower()

    if suffix == ".xls":
        return _load_via_xlrd(path)
    if suffix in (".xlsx", ".xlsm"):
        return _load_via_openpyxl(path)
    raise ValueError(f"不支持的文件格式: {suffix}")


def load_restricted_regions_from_bytes(
    data: bytes, filename: str = "upload.xlsx"
) -> list[dict[str, str | None]]:
    """从内存字节严格解析限发区域。"""
    import tempfile

    suffix = Path(filename).suffix.lower()
    if suffix not in (".xls", ".xlsx", ".xlsm"):
        raise ValueError(f"不支持的文件格式: {suffix}")

    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(data)
        tmp_path = Path(tmp.name)

    try:
        return load_restricted_regions_from_xlsx(tmp_path)
    finally:
        tmp_path.unlink(missing_ok=True)


# --- Helpers ---


def _cell_str(ws, row: int, col: int) -> str:
    """读取单元格值（不处理合并），返回 strip 后的字符串。"""
    value = ws.cell(row, col).value
    if value is None:
        return ""
    return str(value).strip()


def _cell_str_resolved(
    ws, row: int, col: int, merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]]
) -> str:
    """读取单元格值（解析合并到锚点），返回 strip 后的字符串。"""
    anchor = _merged_anchor(merged_lookup, row, col)
    value = ws.cell(*anchor).value
    if value is None:
        return ""
    return str(value).strip()


def _merged_anchor(
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    row: int,
    col: int,
) -> tuple[int, int]:
    bounds = merged_lookup.get((row, col))
    if bounds is None:
        return (row, col)
    return (bounds[0], bounds[1])


def _merged_row_span(
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    row: int,
    col: int,
) -> tuple[int, int]:
    """返回 (start_row, end_row)。"""
    bounds = merged_lookup.get((row, col))
    if bounds is None:
        return (row, row)
    return (bounds[0], bounds[2])


def _is_merged(
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]],
    row: int,
    col: int,
) -> bool:
    return (row, col) in merged_lookup


def _build_merged_lookup(ws) -> dict[tuple[int, int], tuple[int, int, int, int]]:
    """构建合并单元格映射表，并验证合并范围仅限 A 列。"""
    lookup: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    for cell_range in ws.merged_cells.ranges:
        if cell_range.min_col != 1 or cell_range.max_col != 1:
            raise ImportRuleError(
                row=cell_range.min_row,
                column="A",
                reason=f"合并单元格 {cell_range} 跨越了 A 列以外的列",
                suggestion="请确保合并单元格仅在 A 列（产品名称）纵向合并",
            )
        bounds = (
            cell_range.min_row,
            cell_range.min_col,
            cell_range.max_row,
            cell_range.max_col,
        )
        for row_idx in range(cell_range.min_row, cell_range.max_row + 1):
            lookup[(row_idx, 1)] = bounds
    return lookup


def _normalized_city(city: str) -> str | None:
    """过滤备注类城市名，返回 None 表示不限市。"""
    city = city.replace(" ", " ").replace("　", " ").strip()
    if city in _CITY_REMARKS:
        return None
    return city or None


def _split_skus(text: str) -> list[str]:
    """按换行符拆分 SKU。

    Only newlines act as delimiters.
    """
    normalized = text.replace("�", " ").replace("\u3000", " ")
    parts = re.split(r"[\n\r]+", normalized)
    return [p.strip() for p in parts if p.strip()]

def _has_newline(text: str) -> bool:
    return "\n" in text or "\r" in text


def _dedupe_with_province_override(
    rules: list[dict[str, str | None]],
) -> list[dict[str, str | None]]:
    """去重 + 省级限发覆盖市级限发。"""
    seen: set[tuple[str, str, str]] = set()
    unique: list[dict[str, str | None]] = []
    for r in rules:
        key = (r["sku_code"], r["province"], r["city"] or "")
        if key in seen:
            continue
        seen.add(key)
        unique.append(r)

    province_only: set[tuple[str, str]] = set()
    for r in unique:
        if not r["city"]:
            province_only.add((r["sku_code"], r["province"]))

    return [
        r
        for r in unique
        if not r["city"] or (r["sku_code"], r["province"]) not in province_only
    ]


# --- XLSX parser (openpyxl) ---


def _load_via_openpyxl(path: Path) -> list[dict[str, str | None]]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    try:
        max_row = ws.max_row
        if max_row is None or max_row < 2:
            raise ImportRuleError(
                row=1, column="A", reason="Excel 无数据行", suggestion="请添加数据"
            )

        merged_lookup = _build_merged_lookup(ws)

        # 校验表头
        expected = {"产品名称", "省", "市"}
        actual = {
            _cell_str(ws, 1, 1),
            _cell_str(ws, 1, 2),
            _cell_str(ws, 1, 3),
        }
        if not expected.issubset(actual):
            raise ImportRuleError(
                row=1,
                column="A-C",
                reason=f"表头缺失或不符合要求，实际: {actual}",
                suggestion="第一行表头必须包含：产品名称、省、市",
            )

        processed_merged: set[tuple[int, int]] = set()
        rules: list[dict[str, str | None]] = []

        for row_idx in range(2, max_row + 1):
            product_raw = _cell_str(ws, row_idx, 1)
            province = _cell_str(ws, row_idx, 2)
            city_raw = _cell_str(ws, row_idx, 3)

            # 整行全空 → 跳过
            if not product_raw and not province and not city_raw:
                continue

            sku_merged = _is_merged(merged_lookup, row_idx, 1)

            # 省/市列不允许换行符
            if province and _has_newline(province):
                raise ImportRuleError(
                    row=row_idx,
                    column="B",
                    reason=f"单元格包含换行符: {province!r}",
                    suggestion="请拆分为多行",
                )
            if city_raw and _has_newline(city_raw):
                raise ImportRuleError(
                    row=row_idx,
                    column="C",
                    reason=f"单元格包含换行符: {city_raw!r}",
                    suggestion="请拆分为多行",
                )

            # 合并单元格取锚点值，非合并取原始值（换行符作为多 SKU 分隔符是合法的）
            product = _cell_str_resolved(ws, row_idx, 1, merged_lookup) if sku_merged else product_raw

            # 无 SKU 且不在合并范围内 → 报错
            if not product and not sku_merged:
                raise ImportRuleError(
                    row=row_idx,
                    column="A",
                    reason="SKU 为空，且不在 SKU 合并单元格范围内",
                    suggestion="请填写 SKU，或使用真实的 SKU 纵向合并单元格",
                )

            # 合并范围内已处理 → 跳过
            if sku_merged:
                anchor = _merged_anchor(merged_lookup, row_idx, 1)
                if anchor in processed_merged:
                    continue
                processed_merged.add(anchor)
                if not product:
                    raise ImportRuleError(
                        row=anchor[0],
                        column="A",
                        reason="SKU 合并单元格首格为空",
                        suggestion="请在合并区域的第一个单元格填写 SKU",
                    )

            # 省和市都为空 → 报错（合并区域内无省的行自动跳过，不会走到这里）
            if not province and not city_raw:
                raise ImportRuleError(
                    row=row_idx,
                    column="B",
                    reason="省和市均为空，无法生成限发规则",
                    suggestion="请填写省，或同时填写省和市",
                )

            # 省为空但市有值 → 报错
            if not province and city_raw:
                raise ImportRuleError(
                    row=row_idx,
                    column="B",
                    reason="省为空但市/区有值",
                    suggestion="请填写省，或清除市/区的内容",
                )

            # 省为空 → 跳过
            if not province:
                continue

            skus = _split_skus(product)
            if not skus:
                raise ImportRuleError(
                    row=row_idx,
                    column="A",
                    reason="SKU 拆分后为空",
                    suggestion="请填写有效的 SKU",
                )

            city = _normalized_city(city_raw)

            if sku_merged:
                start, end = _merged_row_span(merged_lookup, row_idx, 1)
                for region_row in range(start, end + 1):
                    r_province = _cell_str(ws, region_row, 2)
                    r_city_raw = _cell_str(ws, region_row, 3)
                    if not r_province:
                        continue
                    r_city = _normalized_city(r_city_raw)
                    for sku in skus:
                        rules.append(
                            {"sku_code": sku, "province": r_province, "city": r_city}
                        )
            else:
                for sku in skus:
                    rules.append(
                        {"sku_code": sku, "province": province, "city": city}
                    )

        return _dedupe_with_province_override(rules)
    finally:
        wb.close()


# --- XLS parser (xlrd, legacy) ---


def _load_via_xlrd(path: Path) -> list[dict[str, str | None]]:
    """旧版 .xls 解析，同样执行严格校验。"""
    import xlrd

    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)

    if ws.nrows < 2:
        raise ImportRuleError(
            row=1, column="A", reason="Excel 无数据行", suggestion="请添加数据"
        )

    # 校验表头
    expected = {"产品名称", "省", "市"}
    actual = set()
    for col_idx in range(min(3, ws.ncols)):
        val = ws.cell_value(0, col_idx)
        if val:
            actual.add(str(val).strip())
    if not expected.issubset(actual):
        raise ImportRuleError(
            row=1,
            column="A-C",
            reason=f"表头缺失或不符合要求，实际: {actual}",
            suggestion="第一行表头必须包含：产品名称、省、市",
        )

    rules: list[dict[str, str | None]] = []

    for row_idx in range(1, ws.nrows):
        product = _cell_str_xls(ws, row_idx, 0)
        province = _cell_str_xls(ws, row_idx, 1)
        city_raw = _cell_str_xls(ws, row_idx, 2)

        if not product and not province and not city_raw:
            continue

        if product and _has_newline(product):
            raise ImportRuleError(
                row=row_idx + 1,
                column="A",
                reason=f"单元格包含换行符: {product!r}",
                suggestion="请拆分为多行（使用换行符分隔多个 SKU）",
            )

        if not product:
            raise ImportRuleError(
                row=row_idx + 1,
                column="A",
                reason="SKU 为空",
                suggestion="请填写 SKU",
            )

        if not province and city_raw:
            raise ImportRuleError(
                row=row_idx + 1,
                column="B",
                reason="省为空但市/区有值",
                suggestion="请填写省，或清除市/区的内容",
            )

        if not province:
            continue

        skus = _split_skus(product)
        if not skus:
            raise ImportRuleError(
                row=row_idx + 1,
                column="A",
                reason="SKU 拆分后为空",
                suggestion="请填写有效的 SKU",
            )

        city = _normalized_city(city_raw)
        for sku in skus:
            rules.append({"sku_code": sku, "province": province, "city": city})

    return _dedupe_with_province_override(rules)


def _cell_str_xls(ws, row_idx: int, col_idx: int) -> str:
    try:
        value = ws.cell_value(row_idx, col_idx)
    except IndexError:
        return ""
    if value is None:
        return ""
    return str(value).strip()